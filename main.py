# Jared Tauler


#### Imports ####
import cryptography
import pymysql

# GUI
import PySimpleGUI as sg

# Connecting to DB / managing DB data
import pandas as pd
import regex as re # just for searching

from sqlalchemy import create_engine

from openpyxl.utils.dataframe import dataframe_to_rows as df_to_rows # For exporting from DB server

# PDF
import webbrowser

# E-mail
import smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Logging
import openpyxl as pyxl
from datetime import datetime as dt

import os


#### Functions ####

# Quick window that can be opened and closed while program is doing something that takes time.
class StickyPopup:
	def __init__(self, Text):
		self.window = sg.Window("", layout=[[sg.Text(Text)]], no_titlebar=True, font=FONT, keep_on_top=True)
		self.window.read(timeout=0) # immediately read

	# Kill object and window
	def __del__(self):
		self.window.Close()


# Show quick message.
def Notify(Text, Time):
	# only a function because dont want to carry all these options around.
	sg.PopupTimed(
		Text, auto_close_duration=Time,
		keep_on_top=True, non_blocking=True, font=FONT, no_titlebar=True, button_type=5
	)


# Load Config Files
def LoadConfig():
	if not os.path.exists("config"):
		os.makedirs("config")

	for File in os.listdir("config"):
		# if File not in Exlcude:
		try: exec(open("config/" + File, "r+").read(), globals())
		except: print("Error executing " + File)


# Execute to DB
def ExecuteDB(PreSchema, PostSchema):
	if REMOTEDB:
		C = DB.begin()
		Schema = "`" + SCHEMA + "`."
	else:
		C = DB
		Schema = ""

	C.execute(PreSchema + Schema + PostSchema)


# Update dataframe. Used in like every function.
def dfUpdate(Table):
	if REMOTEDB:
		return pd.read_sql_table(Table, con=DB, schema=SCHEMA)
	else:
		return pd.read_sql_query("SELECT * FROM " + Table, DB)


# Send emails
def SendEmail(Port, Service, Password, Sender, Receiver, Subject, Text, Filename):
	popup = StickyPopup("Sending E-mail...")

	Email = MIMEMultipart()
	Email["From"] = Sender
	Email["To"] = Receiver
	Email["Subject"] = Subject

	if Filename is not None:
		with open(Filename, "rb") as File:
			Attachment = MIMEBase("application", "octet-stream")
			Attachment.set_payload(File.read())

		# Encode file so it can be sent through,
		encoders.encode_base64(Attachment)

		# todo why is attatchment breaking
		Attachment.add_header(
			"Content-Disposition",
			"attachment; filename={Filename}",
		)

	# Attach stuff to email.
		Email.attach(Attachment) # Attachment
	Email.attach(MIMEText(Text, "plain")) # Text

	# Send email.
	with smtplib.SMTP_SSL(Service, Port, context=ssl.create_default_context()) as server:
		server.login(Sender, Password)
		server.sendmail(Sender, Receiver, Email.as_string())

	del popup
	Notify("Mail sent", 1)


# Update table on main window
def TableUpdate(df, window, values):
	try:
		df = df.sort_values(by=HEADING[values["Sort"]]) # Sort by whatever sort box has
	except:
		Notify("Bad sort option", 2)
	# Filter out stuff that doesn't want to be shown
	if SHOWOPTIONS[values["ShowMe"]] == "out":
		df = df.loc[df["status"].isin(["Out", "Late"])]

	# Unnecessary
	# elif SHOWOPTIONS[values["ShowMe"]] == "late":
	# 	df = df.loc[df["status"].isin(["Late"])]

	elif SHOWOPTIONS[values["ShowMe"]] == "in":
		df = df.loc[df["status"].isin([None])]

	elif SHOWOPTIONS[values["ShowMe"]] == "all":
		pass

	df.fillna("", inplace=True)  # Make NULL blank

	# Refine dataframe.
	try:
		df = df[df["name"].str.contains(values["refine"].lower(), flags=re.IGNORECASE)]
	except: # Escape characters will cause an exception.
		pass


	# Go through dataframe, create a list of colors.
	Colors = []
	for i, STATE in enumerate(df["status"].values.tolist()):
		if STATUS.get(STATE) is not None: # Only Color item with not Null status
			Color = STATUS.get(STATE)
			Colors.append((i, Color))

	# Get rid of any unwanted columns. This is done last as it will break checks for stuff if done first.
	df = df[[i for i in HEADING.values()]]

	# I have NO idea why, but if first row gets colored, it will stay that color and wont clear itself after an update
	# like the other rows do. I think it's something with PySG.
	# A workaround is just setting the first row to be the same color as the window color before updating table.
	window.Element("table").Update(row_colors = [(0, window.BackgroundColor)])

	# Update table element. 2nd line searched dataframe for whatever is in refine box.
	window.Element("table").Update(
		values=[
			row for row in df.values.tolist()
		],
		row_colors=Colors
	)

	# Update item counter
	window.Element("LenEntry").Update(
		str(len(window.Element("table").get())) + " items"
	)

	print("Updated Table.")

	# Return a copy of the sorted frame. Only the table double clicking feature needs this.
	return df


def IsItem(Id, df):
	try:
		if df.loc[df["id"].isin([Id])]["quantity"].isnull().item() is False:
			return True
		else:
			return
	except Exception as E:
		sg.popup_error(
			"There was an error while looking for the entry in the database, this is mostly likely"
			" because of bad data in the database.\nID that caused an error was " + Id, E,
			title="Error", font=FONT
		)


# Log events in the log workbook
def Log(df, Event, Id):
	layout = [[sg.Text("Logging...")]]
	window = sg.Window(title="", layout=layout, no_titlebar=True, finalize=True, font=FONT)

	# Open the document.
	ws = wb.active

	AppendList = []

	# Append to excel document:
	if Id is not None:
		AppendList = (
			# Locate row where ID is in the id colum:
			df.loc[df["id"].isin([Id])]
				# Only get columns in LOGHEADING dictionary.
				[list(LOGHEADING.values())]
					# Shave off everything except the data.
					.values.tolist()
						# get first (and only) object in the array:
						[0]
		)

	ws.append(
		# Also add comment about what the log was for
		[Event, dt.now(), ""] + AppendList
	)

	# Save.
	wb.save(WBNAME)
	window.Close()


# Check ID when scanning
def CheckID(Id, IgnoreMissing=False):
	Id = str(Id)

	if Id == "None" or Id.isspace() or Id == "": # check for as none string because converted Id to str
		return None
	else:
		df = dfUpdate("item") # Fetch data from DB

		# check if ID is is a box ID.
		if Id in [i for i in df["box"].tolist()]: # if a tool shares an ID with a box it can never be scanned.
			# While this is a little dumb, this means that barcodes that are already on items can be in the system without
			# give them new barcodes

			# Pick from list of items in box
			Id = BoxPick(
				df.loc[df["box"].isin([Id])],
				PICKSHOW
			)

		# check if ID exists as an item.
		if Id in [i for i in df["id"].tolist()]:
			return Id

		else:
			sg.popup_error("ID not present in database.", font=FONT)
			return None


# def ExceptionLog(): #todo log exceptions


## Connect to DB ##
def ConnectDB(Config):
	def ConfigureDB():
		space = 14
		wide = 15
		remote = [
			[sg.Text("Network address:", size=(space, 0)), sg.InputText(k="HOST", size=(wide, 0))],
			[sg.Text("Username:", size=(space, 0)), sg.InputText(k="USER", size=(wide, 0))],
			[sg.Text("Password:", size=(space, 0)), sg.InputText(k="PASSWORD", size=(wide, 0))],
			[sg.Text("SCHEMA:", size=(space, 0)), sg.InputText(k="SCHEMA", size=(wide, 0))],
			[sg.Checkbox("Use remote database", default=True, k="REMOTEDB", )],
		]
		local = [
			[
				sg.InputText(k="DBLOCATION", size=(17, 0)),
				sg.Button("Select DB File", key="filebrowse")
			]
		]
		Layout = [
			[sg.Frame("Remote DB", layout=remote),],
			[sg.Frame("Local DB", layout=local)],
			[sg.Submit(), sg.Quit()]
		]

		window = sg.Window("Configuration Editor", Layout, finalize=True, font=FONT)

		# Try updating elements with preexisting values. Don't care if exception.
		try: window.Element("HOST").update(HOST)
		except: pass
		try: window.Element("USER").update(USER)
		except: pass
		try: window.Element("PASSWORD").update(PASSWORD)
		except: pass
		try: window.Element("REMOTEDB").update(REMOTEDB)
		except: pass
		try: window.Element("SCHEMA").update(SCHEMA)
		except: pass
		try: window.Element("DBLOCATION").update(DBLOCATION)
		except: pass

		while True:
			event, values = window.read()
			if event == sg.WINDOW_CLOSED or event == "Quit":
				quit()

			elif event == "filebrowse":
				try:
					window.Element("DBLOCATION").update(
						sg.PopupGetFile("Select a database file.")
					)
				except:
					sg.PopupError("Error while getting file.")

			elif event == "Submit":
				Config = open("config/db.txt", "w")

				Config.writelines(
					[
						"# If program is refusing to start, deleting this file might fix it."
						"\nHOST = \"" + values["HOST"] + "\"",
						"\nUSER = \"" + values["USER"] + "\"",
						"\nPASSWORD = \"" + values["PASSWORD"] + "\""
						"\nSCHEMA = \"" + values["SCHEMA"] + "\""
						"\nREMOTEDB = " + str(values["REMOTEDB"]) +
						"\nDBLOCATION = \"" + values["DBLOCATION"] + "\""
					],
				)

				Config.close()
				window.Close()
				break

	def NoConnection(HOST, E):
		Layout = [
			[
				sg.Text(
					"Failed to connect to database at"
				),
				sg.Text(key="HOST", size=(20, 0), relief="solid")
			],
			[
				sg.Text(
					"Is this computer connected to the same network as the database?\n"
					"Is the database login configuration correct?\n"
				)
			],
			[sg.Multiline(size=(60, 5), key="Exception", disabled=True)],  # Text on here can be selected and copied.
			[sg.Button("Retry", key="RETRY"), sg.Button("Reconfigure", key="CONFIG"), sg.Button("Quit", key="QUIT")]
		]

		Window = sg.Window("Connection failure", Layout, finalize=True, font=FONT)
		Window.Element("HOST").update(HOST)
		Window.Element("Exception").update(E)
		event, values = Window.read()

		if event == sg.WINDOW_CLOSED or event == "QUIT":
			quit()

		elif event == "CONFIG":
			Window.Close()
			return True

		elif event == "RETRY":
			Window.Close()

	Connected = False
	NeedRead = True

	## Connect to DB
	# Read DB credentials from a file, and then connects to database.
	while not Connected:
		while NeedRead:
			# Open file, set variables.
			try:
				exec(open("config/db.txt", "r+").read(), globals())
				NeedRead = False

			# Make a file if none
			except Exception as E:
				print(E)

				sg.popup_ok(
					"You need to make a configuration file!",
					"Information regarding the database this program connects to will have to be entered.",
					font=FONT
				)

				ConfigureDB()

		# Go straight to ConfigureDB if just configuring from settings menu.
		if Config:
			ConfigureDB()

		# Open database
		try:
			popup = StickyPopup("Connecting to database...")
			if REMOTEDB:
				DBstr = "mysql+pymysql://" \
						+ USER + \
						":" \
						+ PASSWORD + \
						"@" \
						+ HOST
			else:
				DBstr = "sqlite:///" + DBLOCATION

			global DB
			DB = create_engine(DBstr)
			pd.read_sql_table("item", con=DB)  # todo test connection in a better way?

			del popup
			Connected = True
			return (DB)

		# Failure to connect
		except Exception as E:
			del popup
			if NoConnection(HOST, E):
				ConfigureDB()
				NeedRead = True


## Windows ##
# Error window
def WindowError(Text, E):
	layout = [
		[sg.Text(Text)],
		[sg.Multiline(E, size=(60, 5), disabled=True)],  # Text on here can be selected and copied.
		[sg.Button("Okay")]
	]
	window = sg.Window("Error", layout=layout, font=FONT, modal=True) # Make window
	window.read() # wait for something to happen
	window.Close() # close


# Pick clerk
def ClerkPick():
	df = dfUpdate("person")

	df = df.to_dict()

	# Turn values from dataframe into a dict, where name: id
	NameId = dict(
		zip(
			list(df["name"].values()), list(df["id"].values())
		)
	)
	IdName = dict(zip(NameId.values(), NameId.keys())) # swap key and value for when Id is scanned.

	Layout = [
		[sg.Combo(list(NameId.keys()), enable_events=True, key="combo", readonly=True)],
		[sg.InputText(key="input"), sg.Button("Login", key="login")]
	]

	window = sg.Window("Login", Layout, font=FONT, modal=True, disable_close=True)
	while True:
		event, values = window.read()

		if event == "combo":
			window.Element("input").update(NameId[values["combo"]])

		elif event == "login":
			if values["input"] in NameId.values():
				window.Close()
				Event = "New clerk " + values["input"]
				Log(df, Event, None)

				Notify(values["input"] + " is now the clerk.", 2)
				return IdName[values["input"]]

			else:
				Notify("Bad ID", 2)


# Pick student after scanning an item
def HolderPick():
	df = dfUpdate("person")

	df = df.to_dict()

	# Turn values from dataframe into a dict, where name: id
	NameID = dict(
		zip(
			list(df["name"].values()), list(df["id"].values())
		)
	)

	Layout = [
		[sg.Combo(list(NameID.keys()), enable_events=True, key="combo", readonly=True)],
		[sg.InputText(key="input"), sg.Button("Login", key="login")]
	]

	window = sg.Window("Who's checking out?", Layout, font=FONT)
	while True:
		event, values = window.read()

		if event == sg.WINDOW_CLOSED or event == "QUIT":
			return False

		elif event == "combo":
			window.Element("input").update(NameID[values["combo"]])

		elif event == "login":
			if values["input"] in NameID.values():
				window.Close()
				return {Value: key for key, Value in NameID.items()}[values["input"]]


			else:
				sg.popup("Bad ID")


# Generate layout for main window
def WindowMain():
	Headers = list(HEADING.keys())
	ShowOptions = list(SHOWOPTIONS.keys())

	Layout = [
		[
			sg.Table(
				values=[], headings=Headers, key="table", num_rows=30,
				auto_size_columns=False, bind_return_key=True,
			)
		],
		[
			sg.Frame(title="Sort by", pad=(3,3), layout=[
				[
					sg.InputCombo(values = Headers, pad=(3,3), key = "Sort", default_value="Status", enable_events=True,
					readonly=True),
				]
			]),
			sg.Frame(title="Show only", pad=(3,3), layout=[
				[
					sg.InputCombo(values = ShowOptions, pad = (3, 3), key = "ShowMe", default_value="Checked Out", enable_events=True,
					readonly=True)
				]
			]),
			sg.Frame(title="Scan", pad=(3,3), layout=[
				[
					sg.InputText(key = "ScanBox", size = (10,1), pad=(3,3)),
					sg.Button("Go", key="Scan", pad=(3,3), size = (4,0))
				]
			]),
			sg.Button("Settings", key = "setting"),
			sg.Button("Create ID", key="makeid"),
			sg.Button("Send end of day email", key="email"),
			sg.Button("Modify ID", key="modifyid"),
			sg.Button("New Clerk", key="NewClerk"),
			sg.Button("refresh", key="refresh")
		],
		[
			sg.Text("0 rows", size=(10,0), key="LenEntry"), sg.Text("Start typing to search:  "), sg.InputText(key="refine", enable_events=True)
		]
	]

	return sg.Window("Main", layout=Layout, finalize=True, font = FONT, resizable=True, return_keyboard_events=True,)


# Window to pick from a box
def BoxPick(df, Show):
	layout = [
		[sg.Table(
			values=[row for row in df[Show].values.tolist()],
			headings=[dict([(value, key) for key, value in HEADING.items()])[i] for i in Show],
			key = "table",
			bind_return_key=True
		)],
		[sg.InputText(key="ScanBox"), sg.Button("Scan", key="scan", bind_return_key=True)]
	]

	window = sg.Window("Pick from a list", layout, modal=True, font=FONT)
	while True:
		event, values = window.read()

		if event == sg.WINDOW_CLOSED or event == "QUIT":
			quit()

		elif event == "scan" or event == "table":
			window.Close()
			if event == "table":
				return list(df["id"])[int(values["table"][0])] # List of id column, return whichever row was clicked on table.

			return values["ScanBox"]


# Create ID window
def ModifyID(TrueID, Modify):
	# Get a fresh dataframe
	df = dfUpdate("item")

	def GenLayout():
		# Invert headings list.
		InvHeading = dict([(value, key) for key, value in HEADING.items()])

		# Create a list of existing boxes.
		Boxes = []
		for i in df["box"].tolist():
			if i == "": # Ignore None box.
				pass
			elif i not in Boxes: # Add if not already in.
				Boxes.append(i)

		modify = [
			[
				sg.Column([
					[sg.Text("Quantity:", k="textquantity", size=(8, 0))],
					[sg.Text("Maxi:", k="textmaxi", size=(8, 0))],
				]),
				sg.Column([
					[sg.InputText(k="quantity", size=(4, 0))],
					[sg.InputText(k="maxi", size=(4, 0))],
				]),
			],
			[sg.Checkbox(" Delete when out", k="forget")]
		]

		SDSbuttons = [
			[sg.FileBrowse(enable_events=True, k="browse", file_types=(("PDF Files", "*.pdf"), ("All Files", "*.*")))],
			[sg.Button("Test", k="test")]
		]

		BottomButtons = 		[
				sg.Button("Submit", k="submit"),
				sg.Button("Cancel", k="cancel")
			]
		if Modify:
			BottomButtons.append(sg.Button("delete", k="delete"))

		layout = [
			[
				sg.Column([
					[sg.Text(InvHeading["name"], size=(4,0)), sg.InputText(k="name", size=(20, 1)),],
					[sg.Text(InvHeading["id"], size=(4,0)), sg.InputText(k="id", size=(20, 1))],
					[sg.Text(InvHeading["box"], size=(4,0)), sg.Combo(values=Boxes, k="box", size=(20, 1))],
					# The SDS input box was going to be a multiline, but webbrowser acts really funny for some reason if it
					# gets URL from a multiline. todo new lines causing problems?
					[sg.Text(InvHeading["sds"], size=(4,0)), sg.InputText(k="sds", size=(20, 2)), sg.Column(SDSbuttons)]
				]),

			],
			[
				sg.Frame("Consumable Specific", layout=modify)
			],
			BottomButtons

		]
		return layout

	def Delete(ID):
		ExecuteDB(
			"DELETE FROM ", "`item` WHERE (`id` = '" + ID + "');"
		)

	if Modify:
		Title = "Modifying entry"
	else:
		Title = "Add entry to Database"

	window = sg.Window(Title, GenLayout(), modal=True, finalize=True, font=FONT)


	if Modify:
		UpdateStuff = ["name", "box", "sds", "quantity", "maxi", "id"] # List of
		dfNoNull = df.where(df.notnull(), None) # Replace NULL with 0
		for i in UpdateStuff:
			window.Element(i).update(
				dfNoNull.loc[df["id"].isin([TrueID])][i].item()
			)
		# Set forget to True.
		if dfNoNull.loc[df["id"].isin([TrueID])]["forget"].item() == 1:
			window.Element("forget").update(True)

	while True:
		event, values = window.read()

		if event == sg.WINDOW_CLOSED or event == "cancel":
			window.Close()
			break

		# Delete ID
		elif event == "delete":
			try:
				Delete(TrueID)
				sg.popup_ok("Success")
				window.Close()
				return

			except:
				sg.popup_error("Bad ID")

		# Browse for pdf
		elif event == "browse":
			window.Element("sds").update(values["browse"])

		# Test SDS link
		elif event == "test":
			wb.get("windows-default").open(values["sds"])

		elif event == "submit":
			ID = values["id"]
			NAME = values["name"]
			BOX = values["box"]
			QUANTITY = values["quantity"]
			MAXI = values["maxi"]
			SDS = "NULL"
			FORGET = values["forget"]

			try:
				# Check ID
				if (ID.isspace()) or (ID == ""): # Make sure not blank
					raise ValueError("ID cannot be blank.")

				if ID in df["box"].tolist(): # make sure ID is not already in use as a box.
					raise ValueError("Given ID is already assigned to a box.")

				if not Modify:
					if ID in df["id"].tolist(): # Check if ID already exists
						raise ValueError("ID already in use.")

				ID = "'" + ID + "'"


				# Check NAME
				if (NAME.isspace()) or (NAME == ""): # Make sure not blank
					raise ValueError("Name cannot be blank.")

				NAME = "'" + NAME + "'"


				# Check BOX
				if BOX == "":
					BOX = "NULL"

				if BOX in df["id"].tolist():
					raise ValueError("Given box is already assigned to an ID.")

				BOX = "'" + BOX + "'"


				if values["quantity"] == "": # Non-consumable has passed at this point.
					QUANTITY = "NULL"
					MAXI = "NULL"
					FORGET = 0

				# Check consumable specific stuff.
				else:
					# Check QUANTITY
					if QUANTITY.isnumeric() is False: # Test for integer
						raise ValueError("Quantity must be a number.")

					# Check MAXI
					if MAXI == "": # Check maxi if not blank
						MAXI = 0
					else:
						if MAXI.isnumeric() is False:
							raise ValueError("Maxi must be a number.")

					# Convert
					if FORGET:
						FORGET = 1

					else:
						FORGET = 0

				# Remove old entry if modifying. Easier than updating and also lets ID be changed.
				if Modify:
					Delete(TrueID)

				# Write to DB since every check passed.

				ExecuteDB(
					"INSERT INTO", "`item` (`id`, `name`, `box`, `sds`, `quantity`, `maxi`, `forget`) "
					"VALUES "
					"(" +
						str(ID) + ", " +
						str(NAME) + ", " +
						str(BOX) + ", " +
						str(SDS) + ", " +
						str(QUANTITY) + ", " +
						str(MAXI) + ", " +
						str(FORGET)
					+ ");"
				)
				sg.popup_ok("Success")
				window.Close()
				return

			except Exception as E:
				sg.popup_error(E)


# Window to show when scanning a consumable
def WindowConsumable(df, ID, SendEmail):
	# Show remaining on title.
	def Title(window, df):
		NAME = df.loc[df["id"].isin([ID])]["name"].item()
		QUANTITY = str(int(df.loc[df["id"].isin([ID])]["quantity"].item()))
		MAXI = str(int(df.loc[df["id"].isin([ID])]["maxi"].item()))
		window.Element("title").update(
			"There are " + QUANTITY + "\n" +
			NAME +
			"\nremaining,\n"
			"warning at " + MAXI + "."
		)
		return QUANTITY, MAXI, NAME

	# todo make less ugly
	main = [
		[
			sg.Button("SDS", key="sds", size=(8, 3)),
			sg.Button("Used one", key="used", size=(8, 3)),
			sg.Button("Close", key="close", size=(8, 3)),
			sg.Button("Modify", key="modify", size=(8, 3))

		]
	]
	layout = [
		[
			sg.Column([
				[sg.Text(k="title", size=(20, 0), font=FONT), ]
			], justification="center")
		],
		[
			sg.Frame("", main)
		]
	]
	window = sg.Window("", layout, finalize=True, modal=True, font=FONT)

	## Update window elements.

	QUANTITY, MAXI, NAME = Title(window, df)

	while True:
		event, values = window.read()

		if event == sg.WINDOW_CLOSED or event == "close":
			window.Close()
			return

		elif event == "modify":
			window.Close()
			return df, "modify"

		elif event =="sds":
			try:
				# Open link in webbrowser
				wb.open(df.loc[df["id"].isin([ID])]["sds"].item())
			except:
				pass #todo error log

		elif event == "used":
			# Check if on last item and if forgetting item.
			if int(QUANTITY) == 1 and df.loc[df["id"].isin([ID])]["forget"].item() == 1:
				ExecuteDB(
					"DELETE FROM ", "`item` WHERE (`id` = '" + ID + "');"
				)

				EVENT = "Consumable ran out and deleted"
				window.Close()
				# Dont want to read DB as ID doesnt exist anymore.
				Log(wb, df, EVENT)
				return

			else:
				if int(QUANTITY) == 0:
					sg.popup_error(
						"This will result in a NEGATIVE amount of items!\nYou need to update this items quantity using "
						"the modify button."
					)
					EVENT = "Consumable ran out"

				else:
					ExecuteDB(
						"UPDATE ", "`item`"
						"SET " +
						"`quantity` = '" + str(int(QUANTITY) - 1) + "' " +
						"WHERE `id` = '" + ID + "'"
					)

				df = dfUpdate("item")

				# Update log file after reading DB.
				Event = "Consumable used"
				Log(df, Event, Id)

				# Update window title to reflect change
				QUANTITY, MAXI, NAME = Title(window, df)


				if QUANTITY == MAXI:
					# Log(df, Event, Id)
					SendEmail(
						EMAILPORT, EMAILSERVICE, EMAILPASSWORD, EMAILSENDER, EMAILRECEIVER,
						"Maxi reached",
						"Maxi for " + NAME + " has been reached. (" + str(MAXI) + ")",
						None
					)


# Settings window
def WindowSettings(HEADING, LOGHEADING, df):
	def GenLayout():
		EmailTextSize = 20
		email = [
			[sg.Text("Sender Address:", size=(EmailTextSize,0)),
			sg.InputText(EMAILSENDER, key="sender_address")],

			[sg.Text("Sender Password:", size=(EmailTextSize,0)),
			sg.InputText(EMAILPASSWORD, key="sender_password"),],

			[sg.Text("Recipient Address:", size=(EmailTextSize,0)),
			sg.InputText(EMAILRECEIVER, key="recipient_address")],

			[sg.Text("Email Port:", size=(EmailTextSize,0)),
			sg.InputText(EMAILPORT, key="email_port")],

			[sg.Text("Email Service:", size=(EmailTextSize,0)),
			sg.InputText(EMAILSERVICE, key="email_service")],
			[
				sg.Button("Gmail App Password Page", key="app_password"), sg.Button("Open Gmail", key="gmail"),
				sg.Button("Send test email", k="email_test")
			],
			[sg.Button("Apply", k="email_apply")]
		]

		theme = [
			[sg.InputText(k="lookbox"), sg.Button("Try On", k="lookset")],
			[sg.Text("*Case is ignored."), sg.Button("Preview Colors", key="lookpreview")],
		]

		headings = [
			"Tool", "Status",
		]

		tablecolor = [
			[sg.Text("Preview Table:")],
			[
				sg.Table(values=[" "], headings=headings, k="tabletable", num_rows=8,
						 auto_size_columns=False, enable_events=True),
				sg.DropDown(values="", k="tablecolorbox", size=(10, 0)),
				sg.Button("Pick Color", k="tablecolor")
			],
		]

		# Generate all headings
		tableheadingselements = []
		for i in list(df.columns):
			tableheadingselements.append([sg.InputText(key="box"+i), sg.Text(i, size=(10,0)), sg.Checkbox(text="", k="check"+i)])
		tableheadings = [
			[sg.Frame("", tableheadingselements)],
			[sg.Button("Apply", k="headersapply"), sg.Text("Checkboxes decide if header is shown on log file or not. ")]
		]


		look = [
			[sg.Frame("Theme", theme)],
			[sg.Frame("Table Colors", tablecolor)],
			[sg.Button("Apply", k="lookapply")]

		]

		todb = [
			[sg.Button("Import", k="IEimport")],
			[sg.Button("Export", k="IEexport")],
			[sg.Button("DB Configuration Window", k="IEconfig")]
		]

		layout = [
			[
				sg.TabGroup(
					[
						[sg.Tab(title = "Theme & Table Colors", layout=look)],
						[sg.Tab(title="Table Headings", layout=tableheadings)],
						[sg.Tab(title="Email", layout=email)],
						[sg.Tab(title="Database", layout = todb)]

					]
				)
			]
		]

		return sg.Window("Settings", layout, finalize=True, modal=True, font=FONT)

	# Update window elements.
	def Refresh(window, STATUS):
		# Build a table from STATUS and this list of items. if anything in STATUS changes this wont break.
		TableItem = ["Hammer", "Screwdriver", "Magic Wand", "Banana Peel", "Jack", "Fork"]
		Rows = []
		Colors = []
		for i, j in enumerate(TableItem):
			if i < len(STATUS): # Only append after all statuses are represented.
				Rows.append((j, list(STATUS)[i]))
				Colors.append((i, list(STATUS.values())[i]))
			else:
				Rows.append((j))

		window.Element("tabletable").Update(row_colors=Colors, values=Rows)
		window.Element("tablecolorbox").Update(values=list(STATUS.keys()))

		# Update header boxes.
		for KEY in HEADING.keys():
			window.Element("box" + HEADING.get(KEY)).Update(KEY)

		for KEY in LOGHEADING.keys():
			window.Element("check" + HEADING.get(KEY)).Update(True)

		return Rows # Just returning rows so that when table clicked program knows what's being clicked.

	# Simple function for writing stuff that needs to be remembered to files.
	def Save(Folder, File, Lines):
		# Open file and write to it.
		# Don't care about overwriting variable.
		File = open(Folder + "/" + File, "w+")

		File.writelines(Lines)

		File.close()


	def ToExcel(Folder, Name, Tables):
		wb = pyxl.Workbook()
		for Table in Tables:
			df = dfUpdate("item")
			wb.create_sheet(Table)
			ws = wb[Table]

			rows = df_to_rows(df, index=False)

			for r_idx, row in enumerate(rows, 1):
				for c_idx, value in enumerate(row, 1):
					ws.cell(row=r_idx, column=c_idx, value=value)

		wb.remove_sheet(wb["Sheet"]) # Remove initial sheet

		wb.save(Folder + "/" + Name + ".xlsx")

	# Quick windows for deciding how to import data.
	def TablePick():
		layout = [
			[sg.Text("Is this spreadsheet keeping track of items or people?")],
			[sg.Button("Items", k="item"), sg.Button("People", k="person"), sg.Button("Cancel", k="cancel")],
		]
		window = sg.Window(title="Pick an option", layout=layout, modal=True, font=FONT)
		while True:
			event, values = window.read()

			if event == sg.WINDOW_CLOSED or event == "cancel":
				window.Close()
				return None
			elif event == "item":
				window.Close()
				return "item"
			elif event == "person":
				window.Close()
				return "person"

	def WriteMode():
		layout = [
			[sg.Text("Do you want to add the entries in this file to the database, or replace the database with only data from this spreadhseet?")],
			[sg.Button("Add", k="add"), sg.Button("Replace", k="replace"), sg.Button("Cancel", k="cancel")],
		]
		window = sg.Window(title="Pick an option", layout=layout, modal=True, font=FONT)
		while True:
			event, values = window.read()

			if event == sg.WINDOW_CLOSED or event == "cancel":
				window.Close()
				return None
			elif event == "add":
				window.Close()
				return "append"
			elif event == "replace":
				window.Close()
				return "replace"


	window = GenLayout()
	Rows = Refresh(window, STATUS)

	while True:
		event, values = window.read()

		if event == sg.WINDOW_CLOSED or event == "close":
			return df, "close"

		####################### Look Tab
		# Table Frame #
		# Update color pick box when item on table is clicked
		elif event == "tabletable":
			try:
				Selected = Rows[ # Rows is same list table uses,
					values[event][0] # table event returns row clicked in a list, # todo crashing here?
				][1] # 2nd item in list is status.
				# In case a blank statusless entry is clicked
				if Selected in STATUS.keys():
					window.Element("tablecolorbox").Update(Selected)

			except:
				pass


		elif event == "tablecolor":
			STATUS[values["tablecolorbox"]] = sg.askcolor()[1]
			Refresh(window, STATUS)

		elif event == "lookpreview":
			if sg.popup_ok_cancel(
				"Once you're done looking at the preview, either click one of the many \"OK\" buttons, or do "
				"the Alt+F4 hotkey."
			) == "OK": # Dont .lower() as X button returns None.
				sg.preview_all_look_and_feel_themes()

		elif event == "lookset":
			Theme = values["lookbox"]
			if Theme in [i.lower() for i in sg.theme_list()]: # Check if string in box is a pysimplegui theme
				sg.ChangeLookAndFeel(Theme) # Change if so
				# Remake window.
				window.Close()
				window = GenLayout()
				Refresh(window, STATUS)
			else:
				sg.popup_error("Not a valid theme.")

		elif event == "lookapply":
			Save("config", "table.txt",
				[
					"STATUS = " + str(STATUS),
				]
			)
			Theme = values["lookbox"]
			if Theme in [i.lower() for i in sg.theme_list()]: # make sure theme is valid
				Save("config", "theme.txt",
					[
						"THEME = " + "'" + str(Theme) + "'",
					]
				)

			window.Close()
			return None

		####################### Database Tab
		elif event == "IEconfig":
			window.Close()
			ConnectDB(True) # Do DB connect function

		elif event == "IEimport":
			try:
				try:
					File = pd.read_excel(sg.popup_get_file("Select an Excel file."))
				except: raise ValueError("Couldn't read excel spreadsheet.")

				if REMOTEDB:
					sg.popup("Remote DB importing not implemented.")

				else:
					Table = TablePick()
					if Table is None: raise ValueError("Canceled")

					Mode = WriteMode()
					if Mode is None: raise ValueError("Canceled")

					popup = StickyPopup("Updating database...")
					df.to_sql(Table, con=DB, if_exists=Mode, index=False)
					del popup
					Notify("Success", 2)

			except Exception as E:
				try: del popup
				except: pass
				sg.popup_error(E)


		elif event == "IEexport":
			try:
				Folder = sg.popup_get_folder("Folder to save in?")
				if Folder == "" or Folder.isspace():
					raise ValueError("Canceled.")

				Name = sg.popup_get_text("What to call file?")
				if Name == "" or Name.isspace():
					raise ValueError("Canceled.")

				ToExcel(Folder, Name, ["item", "person"])
			except Exception as E:
				sg.popup_error(E)


		####################### Headers Tab
		elif event == "headersapply":
			# Record heading if not blank
			HEADING = {}
			LOGHEADING = {}
			for i in df.columns:
				if values["box" + i] != "":
					HEADING[values["box" + i]] = i

				if values["check" + i]: # If checkbox is true, add to LOGHEADING.
					LOGHEADING[values["box" + i]] = i

			Save("config", "headers.txt",
				[
					"HEADING = "+ str(HEADING) + "\n" +
					"LOGHEADING = " + str(LOGHEADING)
				 ]
			)

			window.Close()
			return None

		####################### Email Tab
		elif event == "app_password":
			webbrowser.open("https:/myaccount.google.com/apppasswords")

		elif event == "gmail":
			webbrowser.open("https://mail.google.com/mail")

		elif event == "email_test":
			try:
				SendEmail(
					values["email_port"], values["email_service"], values["sender_password"], values["sender_address"],
					values["recipient_address"],
					"Test Email",
					"This is a test email. Today's workbook has been attached.",
					WBNAME
				)
			except Exception as E:
				WindowError("Failed to send Email:", E)

		elif event == "email_apply":
			EmailList = [
				("EMAILPASSWORD", values["sender_password"]),
				("EMAILRECEIVER", values["recipient_address"]),
				("EMAILPORT", values["email_port"]),
				("EMAILSERVICE", values["email_service"]),
				("EMAILSENDER", values["sender_address"])
			]
			# Write to config file.
			Save("config", "email.txt",
				[i[0] + "=\"" + i[1] + "\"\n" for i in EmailList]
			)


def WindowFatal(E):
	Error = str(E) + "\n\n\n# Global variables:\n" + str(globals())
	Time = dt.today()

	File = open("crash/" + Time + ".txt", "w")
	File.writelines(Error)
	File.close()

	layout = [
		[sg.Text("There was an uncaught error, the program must close now. A log file has been created.")],
		[sg.Multiline(Error)],
		[sg.Button("Okay")]
	]
	window = sg.Window("Fatal Error", layout=layout)
	window.read()
	quit()

### Default options ###
# Headings. Key is shown in program while value is internal and in DB.
HEADING = {
	"Tool": "name",
	"ID": "id",
	"Box": "box",
	"Status": "status",
	"Out": "out",
	"In": "in",
	"Student": "holder",
	"Clerk": "clerk",
	"SDS": "sds"
}

# Headers to show on in log file. Default is same as main.
LOGHEADING = HEADING

# Status and colors. These are not user changable.
STATUS = {
	"Out": "darkgreen",
	"Missing": "red",
	"Broken": "gray",
	"Late": "red"
}

# Options for filtering and their internal names
SHOWOPTIONS = {
	"Checked Out": "out",
	"All": "all",
	# "Late": "late",
	"Turned in": "in"
}

EMAILPORT = 465
EMAILSERVICE = None

EMAILPASSWORD = None
EMAILSENDER = None
EMAILRECEIVER = None

USER = None
PASSWORD = None
HOST = None

CLERK = "TEST"
SCHEMA = "jack"
FONT = "Helvetica", 15

THEME = "darkblue3"

LATETIME = 1530

# What columns to show when picking an item from a box
PICKSHOW = [
	"id",
	"name"
]

DBLOCATION = "jack.db.sql"
REMOTEDB = False

#### Main Program ####
# try:
LoadConfig()

sg.ChangeLookAndFeel(THEME) # Set theme

DB = ConnectDB(False) ### Setup DB connection

### Create/load workbook ###
FILETIME = dt.now().strftime("%y-%m-%d")
WBNAME = "logs/" + FILETIME + ".xlsx"
# Load workbook, create new if today's isnt present.
try:
	wb = pyxl.load_workbook(WBNAME)
except:
	wb = pyxl.Workbook()

	ws = wb.active

	# Rows can also be appended
	ws.append(["Event", "Time", ""] + list(LOGHEADING.keys()))

	ws.append(["Log File Created", dt.now()])

	wb.save(WBNAME)

### Main Loop ###

# Open main window #
window = WindowMain()
df = dfUpdate("item") # read from DB for first time window population
event, values = window.read(timeout=0) # Immediately read window
SortedFrame = TableUpdate(df, window, values) # Update table

# Remind user what kinda DB they're using.
Notify("Connected to remote database" if REMOTEDB is True else "Using local database", 2)

# window.bind("<FocusIn>", "FocusIn")
# window.bind("<FocusOut>", "FocusOut")


# CLERK = ClerkPick() # Pick first clerk

LateChecked = False
while True:
	event, values = window.read(timeout = 100)

	if event == sg.WINDOW_CLOSED:
		quit()

	elif event == "NewClerk":
		CLERK = ClerkPick()

	elif event == "email":
		try:
			exec(open("config/email.txt", "r+").read())
			if sg.popup_yes_no("Are you sure you're done for today?") == "Yes":
				Latedf = df.loc[df["status"] == "Late"]  # Get late stuff

				if not Latedf.empty:
					# Make message listing late stuff
					LateStuff = list(zip(Latedf["name"], Latedf["id"]))  # Zip list name and ID

					Text = str(len(LateStuff)) + " entries:\n"
					for i in LateStuff:
						Text = Text + i[0] + " (" + i[1] + "), "
					Text = Text + "are late."

				else:
					Text = "Nothing was late today."

				SendEmail(
					EMAILPORT, EMAILSERVICE, EMAILPASSWORD, EMAILSENDER, EMAILRECEIVER,
					"End of day report",
					Text,
					WBNAME # Pass workbook as attachment.
				)
		except:
			sg.popup_error("No e-mail configuration file found! Go to the settings menu and make one.", title="Error")

	# Access settings window
	elif event == "setting":
		WindowSettings(HEADING, LOGHEADING, df)

		# recreate window
		window.Close()
		LoadConfig()
		DB = ConnectDB(False)
		window = WindowMain()

	# Create ID
	elif event == "makeid":
		Id = ModifyID(None, False)
		df = dfUpdate("item")

		Event = "ID Created"
		Log(df, Event, Id)

	elif event == "modifyid":
		Id = CheckID(sg.popup_get_text("Scan the ID you wish to modify."))

		if Id is not None:
			Id = ModifyID(Id, True)
			df = dfUpdate("item")
			CheckID(Id, True)
			Event = "ID Modified"
			Log(df, Event, Id)

	# update scanbox with ID number of double clicked item on table.
	elif event == "table":
		try:
			# Just reading the table element would be much simpler, but if the ID column isnt displayed it would break.
			# This method won't break, as the table is generated from scratch.
			window.Element("ScanBox").update(
				SortedFrame[
					SortedFrame["name"].str.contains(values["refine"].lower(), flags=re.IGNORECASE)]
					["id"].values.tolist()[values["table"][0]
				]
			)
		except:
			pass

	# Turn in or checkout.
	elif event in ["Scan", "\r", "special 16777220", "special 16777221"]: # Funny strings are to detect enter key being pressed.
		window.Element("ScanBox").update("")

		df = dfUpdate("item")
		Id = CheckID(values["ScanBox"])
		if Id is not None:
			Id = str(Id)
			# If Null quantity is not a consumable. it is VERY important that non consumables have NULL rather than 0 for qauntity
			if IsItem(Id, df):
				WindowConsumable(df, Id, SendEmail)

			else:
				Status = df.loc[df["id"].isin([Id])]["status"].item() # Get status column of item
				ItemName = str(df.loc[df["id"].isin([Id])]["name"].item())
				# If True, something is being checked out.
				if Status is None:
					Holder = HolderPick()
					if Holder: # Only update if holder is true.
						Event = "Checked out"
						ExecuteDB(
							"UPDATE ", "`item`" +
							"SET" +
								"`holder` = '" + Holder + "'," +
								"`out` = CURRENT_TIMESTAMP," +
								"`status` = 'Out', " +
								"`clerk` = '" + CLERK + "'" +
							"WHERE `id` = '" + Id + "'"
						)
						Notify(
							ItemName + " checked out by " + Holder
							, 2
						)

				# Check in
				else:
					Event = "Checked in"
					ExecuteDB(
						"UPDATE ", "`item`" +
						"SET" +
							"`holder` = NULL," +
							"`in` = CURRENT_TIMESTAMP," +
							"`status` = NULL, " +
							"`clerk` = '" + CLERK + "'" +
						"WHERE `id` = '" + Id + "'"
					)
					Notify(
						ItemName + " turned in"
						, 2
					)

				df = dfUpdate("item")  # Data has changed, so read DB.
				Log(df, Event, Id)

	# If it's after late time, check for late tools and and mark them as such, also only check when timing out.
	if event == "__TIMEOUT__" and int(dt.now().strftime("%H%M")) >= 1 or event == "refresh":
		# df = dfUpdate() # dont think updating is necessary
		LateList = df.loc[df["status"] == "Out"].values.tolist()

		# Only if latelist isn't empty
		if len(LateList) > 0:
			for Row in LateList:
				ExecuteDB(
					"UPDATE ", "`item` SET `status` = 'Late' WHERE (`id` = '" + Row[0] + "');"
				)

			df = dfUpdate("item")  # Update df as DB just got modified

			SortedFrame = TableUpdate(df, window, values)

	# Catch keys if a textbox isnt selected, ignore if enter key.
	elif event not in ["\r", "special 16777220", "special 16777221"] and \
		len(event) == 1 and window.FindElementWithFocus().Key not in ["ScanBox", "refine"]: # Don't do if one of the textboxes are in focus.
			window.Element("ScanBox").update(values["ScanBox"] + str(event))

	# Update table
	else: # Don't update table if timing out.
		SortedFrame = TableUpdate(df, window, values)
# except Exception as E: WindowFatal(E)

# todo restart program at midnight cuz log file?
